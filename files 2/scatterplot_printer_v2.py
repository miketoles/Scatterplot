"""
Scatterplot Printer - File Browser Version
Select Excel files directly, update dates, and print

Requirements:
- Python 3.8+
- PyQt5
- openpyxl
- pypiwin32 (optional, for auto-printing)

To run: python scatterplot_printer_v2.py
To build .exe: pyinstaller --onefile --windowed scatterplot_printer_v2.py
"""

import sys
import os
import json
import re
from pathlib import Path
from datetime import datetime, timedelta

# Try to import win32 for printing, but make it optional
try:
    import win32print
    import win32api
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QDateEdit, QListWidget, QListWidgetItem,
    QMessageBox, QProgressDialog, QLineEdit, QFileDialog, QCheckBox,
    QGroupBox, QComboBox
)
from PyQt5.QtCore import Qt, QDate, QThread, pyqtSignal
from PyQt5.QtGui import QFont

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


class PrintWorker(QThread):
    """Background worker thread for printing operations"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(bool, str)
    
    def __init__(self, files, print_date, printer_name, date_cell, test_mode=False):
        super().__init__()
        self.files = files
        self.print_date = print_date
        self.printer_name = printer_name
        self.date_cell = date_cell
        self.test_mode = test_mode
        
    def run(self):
        try:
            total = len(self.files)
            success_count = 0
            
            for i, file_path in enumerate(self.files):
                self.progress.emit(int((i / total) * 100), f"Processing {Path(file_path).name}...")
                
                if self.update_and_print(file_path):
                    success_count += 1
                
            self.progress.emit(100, "Complete!")
            
            if self.test_mode:
                self.finished.emit(
                    True, 
                    f"🧪 TEST MODE: Successfully processed {success_count} of {total} files\n\n"
                    f"⚠️ NO FILES WERE SAVED - Original files are unchanged"
                )
            else:
                self.finished.emit(
                    True, 
                    f"✅ Successfully processed {success_count} of {total} files\n\n"
                    f"Files have been updated and saved with the new date"
                )
            
        except Exception as e:
            self.finished.emit(False, f"Error: {str(e)}")
    
    def update_and_print(self, file_path):
        """Update date in Excel file and print"""
        try:
            # Load workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Try to find and update date in header
            date_updated = False
            
            # Check header sections (left, center, right)
            if hasattr(ws, 'oddHeader'):
                header = ws.oddHeader
                
                # Check left, center, right header sections
                for section_attr in ['left', 'center', 'right']:
                    if hasattr(header, section_attr):
                        section = getattr(header, section_attr)
                        if section and 'Date' in str(section):
                            # Use regex to find and replace ONLY the date, preserving underscores
                            import re
                            
                            # Pattern: "Date" with optional ":" followed by optional underscores, 
                            # then date content, then optional underscores
                            # If no underscores exist, add 3 before and 3 after
                            
                            pattern = r'(Date:?\s*_*)([^_]+?)(_*)'
                            
                            def replace_date(match):
                                leading = match.group(1)
                                trailing = match.group(3)
                                
                                # Ensure "Date:" has a colon (add if missing)
                                if not ':' in leading:
                                    leading = leading.replace('Date', 'Date:', 1)
                                
                                # If no leading underscores after "Date:", add 3
                                if not leading.endswith('_'):
                                    leading += '___'
                                
                                # If no trailing underscores, add 3
                                if not trailing:
                                    trailing = '___'
                                
                                return leading + self.print_date + trailing
                            
                            new_text = re.sub(pattern, replace_date, str(section), flags=re.IGNORECASE)
                            
                            setattr(header, section_attr, new_text)
                            date_updated = True
                            print(f"✅ Updated date in header ({section_attr}) to: {self.print_date}")
                            break
            
            # Fallback: Try the specified cell if header update didn't work
            if not date_updated and self.date_cell:
                # Check if cell contains "Date:" pattern
                cell_value = str(ws[self.date_cell].value) if ws[self.date_cell].value else ""
                
                if 'Date' in cell_value:
                    import re
                    pattern = r'(Date:?\s*_*)([^_]+?)(_*)'
                    
                    def replace_date(match):
                        leading = match.group(1)
                        trailing = match.group(3)
                        
                        # Ensure "Date:" has a colon (add if missing)
                        if not ':' in leading:
                            leading = leading.replace('Date', 'Date:', 1)
                        
                        # If no leading underscores after "Date:", add 3
                        if not leading.endswith('_'):
                            leading += '___'
                        
                        # If no trailing underscores, add 3
                        if not trailing:
                            trailing = '___'
                        
                        return leading + self.print_date + trailing
                    
                    new_value = re.sub(pattern, replace_date, cell_value, flags=re.IGNORECASE)
                    ws[self.date_cell] = new_value
                    print(f"✅ Updated date in cell {self.date_cell} to: {self.print_date}")
                else:
                    # Just set the date directly
                    ws[self.date_cell] = self.print_date
                    print(f"✅ Set cell {self.date_cell} to: {self.print_date}")
            
            # Save ONLY if not in test mode
            if not self.test_mode:
                wb.save(file_path)
                print(f"✅ Saved changes to: {file_path}")
            else:
                print(f"🧪 TEST MODE: Changes NOT saved to: {file_path}")
            
            wb.close()
            
            # Print using available method
            if HAS_WIN32 and self.printer_name:
                win32api.ShellExecute(
                    0,
                    "print",
                    file_path,
                    f'/d:"{self.printer_name}"',
                    ".",
                    0
                )
            else:
                # Fallback: Open the file
                os.startfile(file_path)
                print(f"⚠️ Opened file for manual printing: {file_path}")
            
            return True
            
        except Exception as e:
            print(f"Error updating/printing {file_path}: {e}")
            return False


class MainWindow(QMainWindow):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Scatterplot Printer - File Browser")
        self.setMinimumSize(700, 600)
        
        # Data
        self.selected_files = []
        self.config = {
            'base_path': '',
            'date_cell': 'A1',
            'printer_name': '',
            'test_mode': True  # Default to Test Mode ON
        }
        
        self.init_ui()
        self.load_config()
        
    def init_ui(self):
        # Central widget
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(15)
        
        # Settings Section
        settings_group = QGroupBox("⚙️ Settings")
        settings_layout = QVBoxLayout()
        
        # Base path
        base_path_layout = QHBoxLayout()
        base_path_label = QLabel("Default Browse Location:")
        base_path_label.setStyleSheet("font-weight: bold;")
        settings_layout.addWidget(base_path_label)
        
        self.base_path_input = QLineEdit()
        self.base_path_input.setPlaceholderText("L:\\Behavior Plans")
        base_path_layout.addWidget(self.base_path_input)
        
        browse_base_btn = QPushButton("Browse...")
        browse_base_btn.clicked.connect(self.browse_base_path)
        base_path_layout.addWidget(browse_base_btn)
        
        settings_layout.addLayout(base_path_layout)
        
        # Date cell and printer in one row
        config_row = QHBoxLayout()
        
        # Date cell
        date_cell_label = QLabel("Date Cell:")
        config_row.addWidget(date_cell_label)
        self.date_cell_input = QLineEdit()
        self.date_cell_input.setPlaceholderText("A1")
        self.date_cell_input.setMaximumWidth(60)
        config_row.addWidget(self.date_cell_input)
        
        config_row.addSpacing(20)
        
        # Printer
        printer_label = QLabel("Printer:")
        config_row.addWidget(printer_label)
        self.printer_combo = QComboBox()
        self.printer_combo.setEditable(True)
        self.refresh_printers()
        config_row.addWidget(self.printer_combo)
        
        refresh_printer_btn = QPushButton("🔄")
        refresh_printer_btn.clicked.connect(self.refresh_printers)
        refresh_printer_btn.setMaximumWidth(40)
        config_row.addWidget(refresh_printer_btn)
        
        settings_layout.addLayout(config_row)
        
        # Test mode
        self.test_mode_checkbox = QCheckBox("🧪 Test Mode (Don't save files)")
        self.test_mode_checkbox.setStyleSheet("font-weight: bold; color: #d97706; padding: 8px;")
        settings_layout.addWidget(self.test_mode_checkbox)
        
        # Save settings button
        save_settings_btn = QPushButton("💾 Save Settings")
        save_settings_btn.clicked.connect(self.save_config)
        save_settings_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                padding: 8px;
                font-weight: bold;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1d4ed8;
            }
        """)
        settings_layout.addWidget(save_settings_btn)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
        
        # File Selection Section
        files_group = QGroupBox("📁 Selected Files to Print")
        files_layout = QVBoxLayout()
        
        # Add files button
        add_files_btn = QPushButton("➕ Add Excel Files...")
        add_files_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                padding: 12px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        add_files_btn.clicked.connect(self.add_files)
        files_layout.addWidget(add_files_btn)
        
        # File list
        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QListWidget.ExtendedSelection)
        files_layout.addWidget(self.file_list)
        
        # File list controls
        list_controls = QHBoxLayout()
        
        remove_btn = QPushButton("🗑️ Remove Selected")
        remove_btn.clicked.connect(self.remove_selected_files)
        list_controls.addWidget(remove_btn)
        
        clear_btn = QPushButton("Clear All")
        clear_btn.clicked.connect(self.clear_all_files)
        list_controls.addWidget(clear_btn)
        
        files_layout.addLayout(list_controls)
        
        files_group.setLayout(files_layout)
        layout.addWidget(files_group)
        
        # Print Section
        print_group = QGroupBox("🖨️ Print")
        print_layout = QVBoxLayout()
        
        # Date selection
        date_label = QLabel("Print Date:")
        date_label.setStyleSheet("font-weight: bold;")
        print_layout.addWidget(date_label)
        
        date_row = QHBoxLayout()
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate().addDays(1))
        self.date_edit.setDisplayFormat("MMMM d, yyyy")
        date_row.addWidget(self.date_edit)
        
        today_btn = QPushButton("Today")
        today_btn.clicked.connect(lambda: self.date_edit.setDate(QDate.currentDate()))
        date_row.addWidget(today_btn)
        
        tomorrow_btn = QPushButton("Tomorrow")
        tomorrow_btn.clicked.connect(lambda: self.date_edit.setDate(QDate.currentDate().addDays(1)))
        date_row.addWidget(tomorrow_btn)
        
        print_layout.addLayout(date_row)
        
        # Print button
        self.print_btn = QPushButton("🖨️ Print Selected Files")
        self.print_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                padding: 16px;
                font-size: 16px;
                font-weight: bold;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #059669;
            }
            QPushButton:disabled {
                background-color: #d1d5db;
            }
        """)
        self.print_btn.clicked.connect(self.start_printing)
        print_layout.addWidget(self.print_btn)
        
        print_group.setLayout(print_layout)
        layout.addWidget(print_group)
        
        # Set stylesheet
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f8fafc;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #e2e8f0;
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:hover {
                opacity: 0.9;
            }
        """)
        
        self.update_print_button()
    
    def browse_base_path(self):
        """Browse for base path"""
        path = QFileDialog.getExistingDirectory(self, "Select Base Path")
        if path:
            self.base_path_input.setText(path)
    
    def refresh_printers(self):
        """Refresh the list of available network printers"""
        self.printer_combo.clear()
        
        if not HAS_WIN32:
            self.printer_combo.addItem("Manual entry (pywin32 not installed)")
            self.printer_combo.setEditable(True)
            return
        
        try:
            printers = [printer[2] for printer in win32print.EnumPrinters(win32print.PRINTER_ENUM_NAME)]
            self.printer_combo.addItems(printers)
            
            default_printer = win32print.GetDefaultPrinter()
            index = self.printer_combo.findText(default_printer)
            if index >= 0:
                self.printer_combo.setCurrentIndex(index)
                
        except Exception as e:
            QMessageBox.warning(self, "Printer Error", f"Could not load printers: {e}\n\nYou can type the printer name manually.")
            self.printer_combo.setEditable(True)
    
    def add_files(self):
        """Open file dialog to add Excel files"""
        start_path = self.base_path_input.text() or os.path.expanduser("~")
        
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Excel Files to Print",
            start_path,
            "Excel Files (*.xlsx *.xls);;All Files (*.*)"
        )
        
        if files:
            for file in files:
                if file not in self.selected_files:
                    self.selected_files.append(file)
                    self.file_list.addItem(file)
            
            self.update_print_button()
    
    def remove_selected_files(self):
        """Remove selected files from list"""
        for item in self.file_list.selectedItems():
            file_path = item.text()
            if file_path in self.selected_files:
                self.selected_files.remove(file_path)
            self.file_list.takeItem(self.file_list.row(item))
        
        self.update_print_button()
    
    def clear_all_files(self):
        """Clear all files from list"""
        reply = QMessageBox.question(
            self,
            "Clear All Files",
            "Remove all files from the list?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.selected_files.clear()
            self.file_list.clear()
            self.update_print_button()
    
    def update_print_button(self):
        """Update print button text and state"""
        count = len(self.selected_files)
        self.print_btn.setText(f"🖨️ Print {count} File{'s' if count != 1 else ''}")
        self.print_btn.setEnabled(count > 0)
    
    def start_printing(self):
        """Start the printing process"""
        if not self.selected_files:
            QMessageBox.warning(self, "No Files", "Please add files to print first")
            return
        
        # Get settings
        date_cell = self.date_cell_input.text().strip() or 'A1'
        printer_name = self.printer_combo.currentText()
        test_mode = self.test_mode_checkbox.isChecked()
        
        # Format date as "January 1, 2026"
        date_obj = self.date_edit.date().toPyDate()
        month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        print_date = f"{month_names[date_obj.month - 1]} {date_obj.day}, {date_obj.year}"
        
        # Confirm
        if test_mode:
            reply = QMessageBox.question(
                self,
                "Confirm Test Print",
                f"🧪 TEST MODE ENABLED\n\n"
                f"Process {len(self.selected_files)} file(s) for {print_date}?\n\n"
                f"⚠️ Files will be opened and printed but NOT SAVED.\n"
                f"The original files will remain unchanged.\n\n"
                f"This is safe for testing.",
                QMessageBox.Yes | QMessageBox.No
            )
        else:
            reply = QMessageBox.question(
                self,
                "Confirm Printing",
                f"Process {len(self.selected_files)} file(s) for {print_date}?\n\n"
                f"✅ Files WILL BE SAVED with the updated date.\n\n"
                f"Proceed?",
                QMessageBox.Yes | QMessageBox.No
            )
        
        if reply == QMessageBox.No:
            return
        
        # Create progress dialog
        progress = QProgressDialog("Preparing to print...", "Cancel", 0, 100, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setAutoClose(True)
        progress.setAutoReset(True)
        
        # Start worker thread
        self.worker = PrintWorker(
            self.selected_files,
            print_date,
            printer_name,
            date_cell,
            test_mode
        )
        
        self.worker.progress.connect(lambda val, msg: (progress.setValue(val), progress.setLabelText(msg)))
        self.worker.finished.connect(lambda success, msg: self.print_finished(success, msg, progress))
        
        self.worker.start()
        progress.show()
    
    def print_finished(self, success, message, progress):
        """Handle print completion"""
        progress.close()
        
        if success:
            QMessageBox.information(self, "Print Complete", message)
        else:
            QMessageBox.critical(self, "Print Error", message)
    
    def save_config(self):
        """Save configuration"""
        self.config = {
            'base_path': self.base_path_input.text().strip(),
            'date_cell': self.date_cell_input.text().strip() or 'A1',
            'printer_name': self.printer_combo.currentText(),
            'test_mode': self.test_mode_checkbox.isChecked()
        }
        
        try:
            with open('scatterplot_config_v2.json', 'w') as f:
                json.dump(self.config, f, indent=2)
            
            QMessageBox.information(self, "Settings Saved", "Settings have been saved successfully!")
                
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Could not save settings: {e}")
    
    def load_config(self):
        """Load configuration"""
        try:
            if os.path.exists('scatterplot_config_v2.json'):
                with open('scatterplot_config_v2.json', 'r') as f:
                    self.config = json.load(f)
                
                self.base_path_input.setText(self.config.get('base_path', ''))
                self.date_cell_input.setText(self.config.get('date_cell', 'A1'))
                self.test_mode_checkbox.setChecked(self.config.get('test_mode', False))
                
                printer = self.config.get('printer_name', '')
                index = self.printer_combo.findText(printer)
                if index >= 0:
                    self.printer_combo.setCurrentIndex(index)
                
        except Exception as e:
            print(f"Error loading config: {e}")


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
